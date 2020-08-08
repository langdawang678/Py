import openpyxl

"""
获取表单
1.wb.active 被选中，被激活
2.通过索引，wb.worksheets[索引]
3.通过sheet名字， wb['sheet_name']
"""

# 读取excel文件，读取文件之前一定要关闭该文件
# 得到一个workboot对象，Windows下面的路径有反斜杠
wb = openpyxl.load_workbook(r"excel.xlsx")
print(wb)
# <openpyxl.workbook.workbook.Workbook object at 0x03063A00>

# 不直接去获取_sheets属性，称为私有属性。
print(wb._sheets)  # [<Worksheet "Sheet1">, <Worksheet "Sheet2">, <Worksheet "Sheet3">]
# sheetnames 列表中存储的是字符串，_sheets中存储的是对象

# active 是表示被激活，被选择的sheet
active_sheet = wb.active

# 获取所有表单的正确方式
wb.worksheets

# 获取某一个表单，方式1：索引
sheet = wb.worksheets[0]
# # 方式2：
# sheet = wb.get_sheet_by_name("Sheet1")  # 过时，运行后提示DeprecationWarning
# # 方式3：正规用法
# sheet = wb["Sheet1"]  # 这种方式pycharm 不支持sheet点属性的提示

# 读取单元格，从1开始
cell = sheet.cell(1, 1)
print(cell)  # <Cell 'Sheet1'.A1>
print(cell.value)  # 学生

# 获取某一行
print(sheet[1])  # (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>)
# 获取该行的值
for column in sheet[1]:
    print(column.value)

# 获取某一列
print(sheet["A"])

# 获取多行，!!!含头含尾
print(sheet[1:3])
# ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>))

# 获取所有的数据
total_data = list(sheet.rows)  # sheet.rows的类型<class 'generator'>
print(total_data)
# [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>)]
for row in total_data:
    for cell in row:
        print(cell.value)

# 写入单元格
sheet.cell(1, 3).value = "新写入"
# 保存或者另存为
wb.save(r"excel.xlsx")
# 关闭
wb.close()
